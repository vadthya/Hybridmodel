import openpyxl

def getRowCount(filename, sheetname):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return(sheet.max_row)


def getColumnCount(filename, sheetname):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return(sheet.max_column)


def ReadData(filename, sheetname, rownum, columnum):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return sheet.cell(row = rownum, column = columnum).value

def WriteData(filename, sheetname, rownum, columnum,data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    sheet.cell(row = rownum, column = columnum).value = data
    workbook.save(filename)




